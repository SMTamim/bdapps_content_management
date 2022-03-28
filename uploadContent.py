import requests
import selenium.common.exceptions
from PyQt5.QtWidgets import QMainWindow, QApplication, QLabel, QHBoxLayout, QVBoxLayout, QPushButton, QWidget
from PyQt5 import QtWidgets, QtCore
from PyQt5 import QtGui
import sys, time
import os
from db import DB
from revenue_report import RevenueReport
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import xlsxwriter
import openpyxl, random


def raiseException(err):
    print(err)
    exc_type, exc_obj, exc_tb = sys.exc_info()
    file_name = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
    print(exc_type, file_name, exc_tb.tb_lineno)


database = DB()


class UploadWindow(QWidget):
    def __init__(self, username, password):
        super().__init__()
        self.appContentCountExcelDict = {}
        self.appContentCountOnSiteDict = {}
        self.appLastContentDateFromSiteDict = {}
        self.appContentDictionary = {}
        self.username = username
        self.password = password
        self.working_directory = os.path.abspath(os.getcwd())

        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(":/icons/upload.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.setWindowIcon(icon)

        self.setWindowTitle('Upload Window')

        self.setGeometry(200, 200, 300, 250)
        font = QtGui.QFont('Calibri')
        font.setPointSize(15)
        self.setFont(font)

        vBox = QVBoxLayout()
        self.label = QLabel('Upload Window')
        btn = QPushButton("Upload Content")
        btn.clicked.connect(self.uploadLoadingScreen)
        vBox.addWidget(self.label)
        vBox.addWidget(btn)
        self.setLayout(vBox)

        # Info: Open Excel Sheet Button
        self.openExcel = QtWidgets.QPushButton("Open Excel")
        self.openExcel.setFont(font)
        self.openExcel.setObjectName("openExcel")
        self.openExcel.clicked.connect(self.openExcelFunc)
        vBox.addWidget(self.openExcel)

    # Info: Open Upload loading screen
    def uploadLoadingScreen(self):
        self.new_dialog = QtWidgets.QDialog()
        try:
            font = QtGui.QFont()
            font.setFamily("Calibri")
            font.setPointSize(14)

            self.new_dialog.setWindowTitle("Uploading Content")
            self.new_dialog.setWindowIcon(QtGui.QIcon('icon.png'))
            self.new_dialog.setWindowModality(QtCore.Qt.ApplicationModal)
            self.new_dialog.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)
            self.new_dialog.setFont(font)
            image_name = random.choice(['giphy.gif', 'work-load.gif'])
            if image_name == "giphy.gif":
                self.new_dialog.setFixedSize(325, 450)
            else:
                self.new_dialog.setFixedSize(325, 250)

            self.new_dialog.animation = QtGui.QMovie(
                os.path.join(os.path.abspath(os.getcwd()), 'resources', image_name))
            self.new_dialog.title = QtWidgets.QLabel("Wait while I'm trying to upload content!")
            self.new_dialog.title.setAlignment(QtCore.Qt.AlignCenter)
            self.new_dialog.label_loading = QtWidgets.QLabel(self.new_dialog)
            self.new_dialog.label_loading.setStyleSheet("border: 0px;")
            self.new_dialog.label_loading.setMovie(self.new_dialog.animation)
            vBox = QtWidgets.QVBoxLayout()
            vBox.addWidget(self.new_dialog.title)
            hBox = QtWidgets.QHBoxLayout()
            hBox.addWidget(self.new_dialog.label_loading)
            vBox.addLayout(hBox)
            self.new_dialog.setLayout(vBox)
            self.new_dialog.animation.start()
            self.uploadThread = UploadThread(self.username, self.password)
            self.uploadThread.start()
            self.new_dialog.show()
            self.uploadThread.stop_signal.connect(self.stop_upload_thread)

        except Exception as e:
            raiseException(e)

    # info: Close upload thread
    def stop_upload_thread(self, stop_signal: bool):
        if stop_signal:
            self.uploadThread.stop()
            self.new_dialog.animation.stop()
            self.new_dialog.close()

    # Info: Generate Excel Function
    def generateExcelFunc(self):
        try:
            apps = database.fetch_all_apps(self.username)
            try:
                os.chdir('files')
            except FileNotFoundError:
                os.mkdir('files')
                os.chdir('files')
            wb = xlsxwriter.Workbook(self.username + '.xlsx')
            for app in apps:
                ds = wb.add_worksheet(app[0])
                ds.write(0, 0, 'end')
            wb.close()
            os.chdir('..')
        except Exception as e:
            print(e)

    def has_content_excel(self):
        if self.username + ".xlsx" not in os.listdir(os.path.join(os.path.abspath(os.getcwd()), "files")):
            return False
        else:
            return True

    # Info: Open Excel File
    def openExcelFunc(self):
        try:
            if not self.has_content_excel():
                self.generateExcelFunc()
            os.startfile(os.path.join(os.path.relpath(os.getcwd()), "files", self.username + ".xlsx"))
        except Exception as e:
            print(e)

    def call_the_driver_initializer(self):
        self.wd = RevenueReport(self.username, self.password)
        self.wd.initialize_wd()
        self.wd.enter_login_info()
        self.wd.click_login_btn()
        self.driver = self.wd.driver

    def get_jwt_token(self):
        self.call_the_driver_initializer()
        print("\nGetting App Token!\n")
        print("Waiting for site to load...", end='')
        WebDriverWait(self.driver, 120).until(
            ec.element_to_be_clickable(
                (By.XPATH, "//a[@href='https://user.bdapps.com/soltura']//div[@class='common-modules']"))
        )
        print("Done Waiting")

        self.driver.get('https://user.bdapps.com/provisioning-v2/applications')

        print("Waiting again for site to load...", end='')
        WebDriverWait(self.driver, 120).until(
            ec.element_to_be_clickable((By.CSS_SELECTOR, "input[placeholder='Search Applications']"))
        )
        print("Done Waiting")

        for request in self.driver.requests:
            if request is not None:
                if str(request.headers).find('X-PROVISIONING-JWT:') != -1:
                    app_token = str(request.headers)
                    app_token = app_token[app_token.find('X-PROVISIONING-JWT:'): app_token.find('sec-ch-ua-mobile')]
                    app_token = app_token[19:].replace('\n', '')
                    if app_token is not None:
                        print("Got the JWT token.\n")
                        self.driver.get("https://user.bdapps.com/cas/logout")
                        self.wd.close_wd()
                        return app_token.replace(' ', '')
                    else:
                        print("No token found! Trying again..")

                        self.logout_of_bdapps()
                        self.wd.close_wd()
                        self.get_jwt_token()

    def get_user_app_name(self):
        app_token = self.get_jwt_token()
        self.set_headers(app_token)
        json_data = {
            'name': '',
            'skip': 0,
            'limit': 50,
        }

        response = requests.post('https://user.bdapps.com/prov-api/v2/app/search',
                                 headers=self.headers, json=json_data)
        return response.json()

    def set_headers(self, app_token):
        self.headers = {
            'Host': 'user.bdapps.com',
            'Content-Length': '35',
            'Sec-Ch-Ua': '"Chromium";v="97", " Not;A Brand";v="99"',
            'Accept': 'application/json',
            'X-Provisioning-Jwt': app_token,
            'Sec-Ch-Ua-Mobile': '?0',
            'Content-Type': 'application/json',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36',
            'Sec-Ch-Ua-Platform': '"Windows"',
            'Origin': 'https://user.bdapps.com',
            'Sec-Fetch-Site': 'same-origin',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Referer': 'https://user.bdapps.com/provisioning-v2/applications',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'en-US,en;q=0.9',
            'Connection': 'close',
        }

    # Info: Set the content count dictionary of Excel Sheet
    def set_content_count_dict_from_excel_sheet(self):
        contentWorkbook = openpyxl.load_workbook(os.path.abspath(os.path.join('files', f'{self.username}.xlsx')))
        appsOfContentWorkbook = contentWorkbook.sheetnames

        for contentApp in appsOfContentWorkbook:
            sheet = contentWorkbook[contentApp]
            content_list = []
            if str(sheet.cell(row=1, column=1).value).lower() == "end":
                print(f"No content for {contentApp}")
                self.appContentCountExcelDict.update({contentApp: 0})
            else:
                count = 0
                for i in range(1, 60):
                    if str(sheet.cell(row=i, column=1).value).lower() == "end":
                        break
                    else:
                        content_list.append(str(sheet.cell(row=i, column=1).value))
                        count += 1

                self.appContentDictionary.update({contentApp: content_list})
                self.appContentCountExcelDict.update({contentApp: count})

    # info: Get and Set app content count from bd apps
    def get_previous_content_from_bdapps(self, apps, uploadFlag=False):
        print(uploadFlag)
        self.driver.get("https://user.bdapps.com/soltura/manage/myApplications.html")
        time.sleep(2)
        for app in apps:
            if app[1] == "Lite":
                if uploadFlag and self.appContentCountExcelDict[app[0]] == 0:
                    continue

                self.driver.get(app[2])
                WebDriverWait(self.driver, 60).until(
                    ec.presence_of_element_located((By.TAG_NAME, "textarea"))
                )
                contentTextAreas = self.driver.find_elements(By.TAG_NAME, "textarea")

                contentCount = 0
                for textArea in contentTextAreas:
                    if len(str(textArea.get_attribute('value'))) > 0:
                        contentCount += 1
                    else:
                        break
                if contentCount == 0:
                    last_content_date = "No Last Content"
                else:
                    last_content_date = self.driver.find_element(By.ID, f'contentGroupId_{contentCount-1}')
                    last_content_date = last_content_date.find_element(By.TAG_NAME, 'span').get_attribute('innerText')
                print(last_content_date)
                self.appLastContentDateFromSiteDict.update({app[0]: last_content_date})
                self.appContentCountOnSiteDict.update({app[0]: contentCount})

                print(f"{app[0]} has {contentCount} previous content!")

                try:
                    if uploadFlag:
                        self.driver.execute_script("for(j=0; j<59; j++)show();")
                        WebDriverWait(self.driver, 60).until(
                            ec.visibility_of_all_elements_located((By.TAG_NAME, "textarea"))
                        )
                        emptyContentFieldIndex = int(self.appContentCountOnSiteDict[app[0]])
                        totalEmptyContentField = 60 - emptyContentFieldIndex

                        if totalEmptyContentField >= int(self.appContentCountExcelDict[app[0]]):
                            totalEmptyContentField = int(self.appContentCountExcelDict[app[0]])
                        else:
                            print(f"Skipping Content of {app[0]} as it exceeds total supported content!")
                            continue

                        actual_content_index = 0
                        for contentTextArea in range(emptyContentFieldIndex,
                                                     emptyContentFieldIndex + totalEmptyContentField):
                            contentTextAreas[contentTextArea].send_keys(
                                self.appContentDictionary[app[0]][actual_content_index])
                            actual_content_index += 1

                        WebDriverWait(self.driver, 60).until(
                            ec.element_to_be_clickable((By.ID, 'send_msg'))
                        )
                        self.driver.find_element(By.ID, 'send_msg').click()
                        print("Waiting 2 Seconds...")
                        time.sleep(2)
                except Exception as e:
                    raiseException(e)

    def upload_content(self, should_upload=True):
        # Info: Fetch all apps from DB for the respective account
        apps = database.fetch_all_apps(self.username)
        if not self.has_content_excel():        # info: If content excel is not generated generate now
            self.generateExcelFunc()

        self.set_content_count_dict_from_excel_sheet()
        print(self.appContentCountExcelDict)
        try:
            if not should_upload:
                does_have_all_link = self.get_content_add_link(False)
            else:
                does_have_all_link = self.get_content_add_link(True)
            # Info: checks if driver is initialized by the method
            if does_have_all_link:
                # info: if not initialize now
                self.call_the_driver_initializer()
            if should_upload:
                self.get_previous_content_from_bdapps(apps, True)
            else:
                self.get_previous_content_from_bdapps(apps)
                print(self.appContentCountOnSiteDict)
                return self.appLastContentDateFromSiteDict
        except Exception as e:
            raiseException(e)

    def get_subscriber(self, use_url):
        print(use_url)
        use_url = use_url.split('/')[-1].replace('.html', '')
        print(use_url)
        view_url = f"https://user.bdapps.com/soltura/reports/viewReport/{use_url}/1.html"
        self.driver.get(view_url)
        try:
            WebDriverWait(self.driver, 120).until(
                ec.presence_of_element_located((By.CSS_SELECTOR, "div[align='right']"))
            )
            subscriber = self.driver.find_element(By.CSS_SELECTOR, "div[align='right']").get_attribute("innerText")
            print(subscriber)
            return subscriber
        except selenium.common.exceptions.TimeoutException:
            return 0

    def get_content_add_link(self, is_running_from_upload_window=True):
        try:
            if is_running_from_upload_window:
                self.label.setText('Checking if app has link in DB!')
            apps = database.fetch_all_apps(self.username)
            all_app_has_link_in_db = True

            for app in apps:
                if app[2] == '' and app[1] == "Lite":
                    all_app_has_link_in_db = False
                    break
            if not all_app_has_link_in_db:
                self.label.setText("Getting links for all apps in DB!")
                self.call_the_driver_initializer()
                self.driver.get("https://user.bdapps.com/soltura/manage/myApplications.html")
                time.sleep(2)
                self.driver.get("https://user.bdapps.com/soltura/manage/myApplications.html")

                WebDriverWait(self.driver, 60).until(
                    ec.presence_of_all_elements_located((By.CLASS_NAME, "my_app_link"))
                )

                names = self.driver.find_elements(By.CLASS_NAME, "bold_text")
                links = self.driver.find_elements(By.CLASS_NAME, "my_app_link")
                print("Updating Links!")
                try:
                    for i in range(0, len(links)):
                        links[i] = links[i].find_element(By.TAG_NAME, "a").get_attribute('href')
                        print(names[i].get_attribute('innerText'), links[i])
                        print(database.update_app_link((names[i].get_attribute('innerText'), links[i])))
                    print("Closing Driver in get_content_add_link.")
                    self.driver.close()
                    time.sleep(2)
                    return True
                except Exception as e:
                    print(e)
                    return False
            else:
                print("here", all_app_has_link_in_db)
                print("Links were already added.")
                return True
        except Exception as e:
            raiseException(e)
        # self.wd.close_wd()


class UploadThread(QtCore.QThread):
    stop_signal = QtCore.pyqtSignal(bool)

    def __init__(self, username, password):
        super(UploadThread, self).__init__()
        self.uploader = UploadWindow(username, password)

    def run(self) -> None:
        self.uploader.upload_content(True)
        self.stop_signal.emit(True)

    def stop(self):
        self.terminate()


class RefreshTableThread(QtCore.QThread):
    stop_signal = QtCore.pyqtSignal(bool)
    current_working_app_signal = QtCore.pyqtSignal(str)

    def __init__(self, username, password, apps):
        super(RefreshTableThread, self).__init__()
        self.apps = apps
        self.siteObj = UploadWindow(username, password)

    def run(self):
        self.stop_signal.emit(False)
        appLastContentDateFromSiteDict = self.siteObj.upload_content(should_upload=False)
        for app_ in self.apps:
            if app_[1] != "Lite":
                continue
            self.current_working_app_signal.emit(app_[0])
            data = tuple(['last_content_date', appLastContentDateFromSiteDict[app_[0]]])
            data = {app_[0]: data}
            database.update_app_detail(data)
            subscriber_count = self.siteObj.get_subscriber(app_[2])
            data = tuple(['subscriber_count', subscriber_count])
            data = {app_[0]: data}
            database.update_app_detail(data)

        self.siteObj.driver.get("https://user.bdapps.com/cas/logout")
        time.sleep(1)
        self.siteObj.wd.close_wd()

        self.stop_signal.emit(True)

    def stop(self):
        self.terminate()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = UploadWindow("darkslave0", "Tamim@646")
    win.show()
    sys.exit(app.exec_())
